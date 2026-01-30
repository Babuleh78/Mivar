import pandas as pd
import re
from sympy import symbols, solve, Eq, sympify, sqrt, root
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class FormulaCreator:
    def solve_for_all_variables(self, formula_str):
        if not formula_str or not isinstance(formula_str, str):
            return []
        
        formula_str = self.convert_roots_to_power(formula_str)
        
        formula_str_for_sympy = formula_str.replace(" ", "").replace("^", "**")
        variables = set(re.findall(r'[A-Za-z_][A-Za-z0-9_]*', formula_str_for_sympy))
        
        if '=' not in formula_str_for_sympy:
            return []
        
        left_str, right_str = formula_str_for_sympy.split('=', 1)
        
        # Создаем символы для всех переменных
        syms = symbols(' '.join(variables))
        var_dict = {str(sym): sym for sym in syms}
        
        try:
            left_expr = sympify(left_str, locals=var_dict)
            right_expr = sympify(right_str, locals=var_dict)
            equation = Eq(left_expr, right_expr)
            
            results = []
            for var in variables:
                try:
                    solution = solve(equation, var_dict[var])
                    if solution and  len(solution) == 1:
                        for sol in solution:
                            if sol == right_expr:
                                continue

                            result_str = str(sol).replace("**", "^")
                            result_str = self.convert_roots_to_power(result_str)
                            results.append(f"{var} = {result_str}")
                except:
                    continue
            
            return results
        except Exception as e:
            print(f"Ошибка при обработке формулы '{formula_str}': {e}")
            return []
    
    def convert_roots_to_power(self, expr_str):
        
        expr_str = self.convert_sqrt_to_power(expr_str)
        
        expr_str = self.convert_root_to_power(expr_str)
        
        return expr_str
    
    def convert_sqrt_to_power(self, expr_str):
        result = expr_str
        
        # Функция для рекурсивного поиска и замены sqrt
        def replace_sqrt_recursive(text):
            # Ищем sqrt(
            start = text.find('sqrt(')
            if start == -1:
                return text
            
            # Находим соответствующую закрывающую скобку
            depth = 0
            pos = start + 5  # пропускаем "sqrt("
            while pos < len(text):
                if text[pos] == '(':
                    depth += 1
                elif text[pos] == ')':
                    if depth == 0:
                        # Нашли закрывающую скобку
                        inner = text[start+5:pos]
                        
                        # Проверяем, нужно ли оборачивать в скобки
                        if '(' in inner or '+' in inner or '-' in inner or '*' in inner or '/' in inner or '^' in inner:
                            replacement = f'({inner})^(1/2)'
                        else:
                            replacement = f'{inner}^(1/2)'
                        
                        # Заменяем и продолжаем рекурсивно
                        new_text = text[:start] + replacement + text[pos+1:]
                        return replace_sqrt_recursive(new_text)
                    depth -= 1
                pos += 1
            
            return text
        
        # Заменяем все sqrt
        while 'sqrt(' in result:
            new_result = replace_sqrt_recursive(result)
            if new_result == result:
                break
            result = new_result
        
        return result
    
    def convert_root_to_power(self, expr_str):
        result = expr_str
        
        # Функция для рекурсивного поиска и замены root
        def replace_root_recursive(text):
            # Ищем root(
            start = text.find('root(')
            if start == -1:
                return text
            
            # Находим соответствующую закрывающую скобку
            depth = 0
            pos = start + 5  # пропускаем "root("
            comma_pos = -1
            
            while pos < len(text):
                if text[pos] == '(':
                    depth += 1
                elif text[pos] == ')':
                    if depth == 0:
                        # Нашли закрывающую скобку
                        if comma_pos != -1:
                            # Извлекаем аргументы
                            inner = text[start+5:pos]
                            # Разделяем по запятой на первом уровне
                            arg_depth = 0
                            comma_idx = -1
                            
                            for i, ch in enumerate(inner):
                                if ch == '(':
                                    arg_depth += 1
                                elif ch == ')':
                                    arg_depth -= 1
                                elif ch == ',' and arg_depth == 0:
                                    comma_idx = i
                                    break
                            
                            if comma_idx != -1:
                                expr = inner[:comma_idx].strip()
                                n_str = inner[comma_idx+1:].strip()
                                
                                # Проверяем, нужно ли оборачивать в скобки
                                if '(' in expr or '+' in expr or '-' in expr or '*' in expr or '/' in expr or '^' in expr:
                                    replacement = f'({expr})^(1/{n_str})'
                                else:
                                    replacement = f'{expr}^(1/{n_str})'
                                
                                # Заменяем и продолжаем рекурсивно
                                new_text = text[:start] + replacement + text[pos+1:]
                                return replace_root_recursive(new_text)
                        break
                    depth -= 1
                elif text[pos] == ',' and depth == 0:
                    comma_pos = pos
                pos += 1
            
            return text
        
        # Заменяем все root
        while 'root(' in result:
            new_result = replace_root_recursive(result)
            if new_result == result:
                break
            result = new_result
        
        return result
    
    def process_excel_by_column(self, input_file, output_file=None):
        if output_file is None:
            output_file = input_file.replace('.xlsx', '_MIVAR.xlsx')
        
        wb = load_workbook(input_file)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column
        
        for col in range(1, max_col + 1):
            col_data = []
            col_letter = get_column_letter(col)
            
            # Собираем все данные столбца
            for row in range(1, max_row + 1):
                cell_value = ws[f"{col_letter}{row}"].value
                if cell_value:
                    col_data.append(str(cell_value))
                else:
                    col_data.append(None)
            
            processed_data = []
            i = 0
            
            # Обрабатываем данные до "Дополнительно"
            while i < len(col_data) and col_data[i] != "Дополнительно":
                item = col_data[i]
                if item and '=' in item:
                    # Заменяем sqrt и root в исходной формуле
                    item_with_power = self.convert_roots_to_power(item)
                    processed_data.append(item_with_power)
                    
                    transformations = self.solve_for_all_variables(item)
                    processed_data.extend(transformations)
                else:
                    processed_data.append(item)
                i += 1
            
            # Добавляем "Дополнительно" и все строки после него
            while i < len(col_data):
                if col_data[i]:
                    processed_data.append(col_data[i])
                i += 1
            
            # Записываем обработанные данные обратно в столбец
            for idx, value in enumerate(processed_data, start=1):
                if idx <= len(processed_data):
                    ws[f"{col_letter}{idx}"].value = value
                else:
                    # Очищаем ячейки, если данных меньше чем было
                    ws[f"{col_letter}{idx}"].value = None
        
        wb.save(output_file)
        print(f"Результат сохранен в: {output_file}")
        return output_file
